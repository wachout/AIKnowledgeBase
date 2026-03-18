# -*- coding:utf-8 _*-


import os
import json
import base64
# import mammoth
import pandas as pd  # type: ignore
import re
import csv
import logging
import threading
from datetime import datetime

from openpyxl import load_workbook  # type: ignore

 # 使用urllib下载文件
import urllib.request
import urllib.error

# from Files.read_doc import cSingleDoc
# from Files.mineru_pdf import cSinglePdf

from Utils import utils

# 创建线程安全的logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)  # 设置日志级别
logger_lock = threading.Lock()

class CControl():
    
    def __init__(self):
        self.file_path = "conf/tmp/"
        if not os.path.exists(self.file_path):
            os.makedirs(self.file_path)
    
    def save_file(self, file_path, content):
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return file_path
    
    def delete_file(self, file_path, file_name):
        file_path = os.path.join(file_path, file_name)
        if os.path.exists(file_path):
            os.remove(file_path)
            return True
        return False
    
    def read_url(self, _url):
        file_extension = os.path.splitext(_url)[1].lower()
        if file_extension in [".doc", ".docx", ".pdf", ".xls", ".xlsx"]:
            api = "/api/read_file_to_base64"
            param = {"url": _url}
            content = utils.request(param, api)
            _json = json.loads(content)
            result_str = _json["result"]
        
            if result_str.startswith("b'") and result_str.endswith("'") and len(result_str) > 4:
                base64_data = result_str[2:-1]  # 去掉引号
                query = base64.b64decode(base64_data)
                query = query.decode('utf-8')
            else:
                query = ""
            
            content_str = _json["content"]
            if content_str.startswith("b'") and content_str.endswith("'") and len(content_str) > 4:
                base64_data = content_str[2:-1]  # 去掉引号
                content = base64.b64decode(base64_data)
                content = content.decode('utf-8')
            else:
                content = ""
                
            return query, content
        else:
            return None, None
        
    def read_stream_file(self, file_path):
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension in [".doc", ".docx", ".pdf", ".xls", ".xlsx"]:
            
            # 读取文件内容
            file_content = None
            # file_name = "unknown"
            
            file_name = os.path.basename(file_path)
            
            # 如果是文件对象，则读取其内容
            if hasattr(file_path, 'read'):
                file_content = file_path.read()
            elif isinstance(file_path, str) and os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    file_content = f.read()
            else:
                # 文件不存在，返回None, None
                return None, None

            # 对读取的二进制文件进行base64编码
            base64_content = base64.b64encode(file_content).decode('utf-8')

            # 构造包含base64编码数据的请求体
            param = {
                'file_name': file_name,
                'file_content_base64': base64_content,
                'file_extension': file_extension
            }

            api = "/api/read_base64_file_to_base64"
            # param = {"data": data}
            content = utils.request(param, api)
            # print(content)
            _json = json.loads(content)
            result_str = _json["result"]
        
            if result_str.startswith("b'") and result_str.endswith("'") and len(result_str) > 4:
                base64_data = result_str[2:-1]
                query = base64.b64decode(base64_data)
                query = query.decode('utf-8')
            else:
                query = ""
            
            content_str = _json["content"]
            if content_str.startswith("b'") and content_str.endswith("'") and len(content_str) > 4:
                base64_data = content_str[2:-1]  # 去掉引号
                content = base64.b64decode(base64_data)
                content = content.decode('utf-8')
            else:
                content = ""
                
            return query, content
        else:
            return None, None

    def read_csv(self, file_path):
        """
        读取 CSV 文件并生成用于 RAG 的文本描述，同时创建 SQLite 数据库支持文本转 SQL。
        
        - 调用 CsvFileAgent，综合 sheet 描述与列名称/示例值
        - 返回一段可直接写入向量库/检索系统的文本说明
        - 创建 SQLite 数据库（{文件名}_data.db），表的列名对应 CSV 列名
        - 在 SQLite 中保存 _table_metadata 和 _column_metadata 元数据表
        
        Args:
            file_path: CSV 文件路径
        
        Returns:
            str: CSV 的文本描述（若解析失败返回空字符串）
        """
        if not file_path or not isinstance(file_path, str) or not os.path.isfile(file_path):
            return ""
        # 延迟导入，避免在不使用 CSV 功能时增加启动开销
        try:
            from Agent.csv_file_agent import CsvFileAgent
        except Exception:
            return ""
        try:
            agent = CsvFileAgent()
            meta = agent.parse_csv_file(file_path)
            text_summary = meta.get("text_summary", "") or ""
            
            # 创建 SQLite 数据库，支持文本转 SQL 查询
            self._create_csv_sqlite(file_path, meta)
            
            return text_summary
        except Exception:
            return ""
    
    def _create_csv_sqlite(self, file_path, meta):
        """
        为 CSV 文件创建 SQLite 数据库，表的列名对应 CSV 列名。
        
        Args:
            file_path: CSV 文件路径
            meta: CsvFileAgent 解析的元数据
        """
        try:
            import sqlite3
            import pandas as pd
            
            # 基础信息
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            safe_base = re.sub(r"[^\w\u4e00-\u9fa5]", "_", base_name)[:80] or "csv"
            out_dir = os.path.dirname(os.path.abspath(file_path))
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 创建 SQLite 数据库
            sqlite_path = os.path.join(out_dir, f"{safe_base}_data.db")
            conn = sqlite3.connect(sqlite_path)
            cursor = conn.cursor()
            
            # 创建表格元数据表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS _table_metadata (
                    table_name TEXT PRIMARY KEY,
                    sheet_name TEXT,
                    description TEXT,
                    row_count INTEGER,
                    column_count INTEGER,
                    csv_path TEXT,
                    created_at TEXT
                )
            """)
            
            # 创建列元数据表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS _column_metadata (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT,
                    column_name TEXT,
                    column_index INTEGER,
                    data_type TEXT,
                    description TEXT,
                    sample_values TEXT,
                    UNIQUE(table_name, column_name)
                )
            """)
            
            # 读取 CSV 并保存到 SQLite，列名对应 CSV 列名
            df = pd.read_csv(file_path, encoding="utf-8-sig")
            table_name = safe_base
            df.to_sql(table_name, conn, if_exists='replace', index=False)
            
            # 从 meta 获取信息
            sheet_name = meta.get("sheet_name") or base_name
            sheet_desc = meta.get("sheet_description") or ""
            columns = meta.get("columns") or []
            
            # 保存表格元数据
            cursor.execute("""
                INSERT OR REPLACE INTO _table_metadata 
                (table_name, sheet_name, description, row_count, column_count, csv_path, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                table_name,
                sheet_name,
                sheet_desc,
                len(df),
                len(df.columns),
                os.path.abspath(file_path),
                ts
            ))
            
            # 保存列元数据
            for col_idx, col in enumerate(columns):
                cursor.execute("""
                    INSERT OR REPLACE INTO _column_metadata 
                    (table_name, column_name, column_index, data_type, description, sample_values)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    table_name,
                    col.get('name', ''),
                    col_idx,
                    col.get('dtype', 'TEXT'),
                    col.get('description', ''),
                    json.dumps(col.get('sample_values', []), ensure_ascii=False)
                ))
            
            conn.commit()
            conn.close()
            logger.info(f"CSV SQLite 数据库已创建: {sqlite_path}")
        except Exception as e:
            logger.warning(f"创建 CSV SQLite 数据库失败: {e}")


    def split_excel_to_csv(self, file_path):
        """
        将Excel文件拆分为多个CSV文件，并为每个sheet生成描述信息。
        
        - 每个sheet保存为一个CSV文件：{原文件名}__{sheet名}.csv
        - 同名生成一个元数据文件：{CSV文件名}.meta.json，包含：
          - sheet_name: sheet名称
          - sheet_description: sheet描述（目前使用sheet名）
          - columns: 列信息（name、description、dtype等），供后续RAG/CSV智能体使用
        - 创建 SQLite 数据库（{原文件名}_data.db），每个 CSV 保存为一个表，支持文本转 SQL 查询
        - 在 SQLite 中保存 _table_metadata 和 _column_metadata 元数据表
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            list: 包含所有CSV文件路径的列表（元数据写在 .meta.json 中）；
                  同时在同目录生成一份整本 Excel 的「层级目录」Markdown，结构为：
                  Excel 主目录 -> Sheet 子目录 -> 列小目录 -> sample_values 示例。
                  同时生成 SQLite 数据库文件用于文本转 SQL 查询。
        """

        # 延迟导入依赖，避免无必要的启动开销
        try:
            import pandas as pd  # type: ignore
        except Exception:
            pd = None  # type: ignore
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            load_workbook = None  # type: ignore
        try:
            import sqlite3
        except Exception:
            sqlite3 = None  # type: ignore

        if not file_path or not isinstance(file_path, str) or not os.path.isfile(file_path):
            return []

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in [".xls", ".xlsx"]:
            return []

        base_name = os.path.splitext(os.path.basename(file_path))[0]
        safe_base = re.sub(r"[^\w\u4e00-\u9fa5]", "_", base_name)[:50] or "excel"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 将拆分后的 CSV 放在原文件所在目录，便于前端/知识库直接访问
        logger.info(f"split_excel_to_csv 源文件路径: {file_path}")
        dir_path, _ = os.path.split(file_path)
        out_dir = dir_path or "."
        logger.info(f"split_excel_to_csv 输出目录: {out_dir}")

        os.makedirs(out_dir, exist_ok=True)

        output_files = []
        sqlite_conn = None
        sqlite_path = ""
        
        # 创建 SQLite 数据库用于文本转 SQL 查询
        if sqlite3 is not None:
            try:
                sqlite_path = os.path.join(out_dir, f"{safe_base}_data.db")
                sqlite_conn = sqlite3.connect(sqlite_path)
                cursor = sqlite_conn.cursor()
                
                # 创建表格元数据表
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS _table_metadata (
                        table_name TEXT PRIMARY KEY,
                        sheet_name TEXT,
                        description TEXT,
                        row_count INTEGER,
                        column_count INTEGER,
                        csv_path TEXT,
                        created_at TEXT
                    )
                """)
                
                # 创建列元数据表
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS _column_metadata (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        table_name TEXT,
                        column_name TEXT,
                        column_index INTEGER,
                        data_type TEXT,
                        description TEXT,
                        sample_values TEXT,
                        UNIQUE(table_name, column_name)
                    )
                """)
                sqlite_conn.commit()
                logger.info(f"创建 SQLite 数据库: {sqlite_path}")
            except Exception as e:
                logger.warning(f"创建 SQLite 数据库失败: {e}")
                sqlite_conn = None
        
        catalog_lines = [
            f"# Excel 文件目录：{base_name}\n",
            f"- 源文件路径: {os.path.abspath(file_path)}\n",
            f"- 拆分时间: {ts}\n\n",
        ]

        # 方案A：优先使用 pandas（支持 xls/xlsx、多 sheet），并输出 sheet/列的描述元数据与整本目录
        if pd is not None:
            try:
                # sheet_name=None 读取所有 sheet，返回 dict[sheet_name] = DataFrame
                sheets = pd.read_excel(file_path, sheet_name=None)
                catalog_lines.append(f"- Sheet 总数: {len(sheets or {})}\n\n")
                for sheet_name, df in (sheets or {}).items():
                    safe_sheet = re.sub(r"[^\w\u4e00-\u9fa5]", "_", str(sheet_name))[:60] or "Sheet"
                    csv_path = os.path.join(out_dir, f"{safe_base}__{safe_sheet}.csv")
                    # utf-8-sig 兼容 Excel 打开 CSV 的中文显示
                    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
                    # 生成元数据：sheet描述 + 列描述（目前用列名+类型+示例值作为描述）
                    meta = {
                        "sheet_name": str(sheet_name),
                        "sheet_description": f"工作表「{sheet_name}」，包含 {len(df.columns)} 列，{len(df)} 行数据。",
                        "columns": []
                    }
                    catalog_lines.append(f"## Sheet: {sheet_name}\n\n")
                    catalog_lines.append(f"- 描述: {meta['sheet_description']}\n")
                    catalog_lines.append(f"- 对应 CSV: {os.path.abspath(csv_path)}\n\n")
                    catalog_lines.append("### 列信息\n\n")
                    for col in df.columns:
                        col_series = df[col]
                        # 取前若干非空示例值
                        sample_values = []
                        for v in col_series:
                            if pd.isna(v):
                                continue
                            sample_values.append(str(v))
                            if len(sample_values) >= 5:
                                break
                        col_meta = {
                            "name": str(col),
                            "dtype": str(col_series.dtype),
                            "description": f"列「{col}」，数据类型 {col_series.dtype}，示例值: {', '.join(sample_values[:3])}" if sample_values else f"列「{col}」，数据类型 {col_series.dtype}。",
                            "sample_values": sample_values,
                        }
                        meta["columns"].append(col_meta)
                        samples_show = f" 示例值: {', '.join(sample_values[:3])}" if sample_values else ""
                        catalog_lines.append(f"- {col_meta['name']} ({col_meta['dtype']}): {col_meta['description']}{samples_show}\n")
                    catalog_lines.append("\n")
                    meta_path = csv_path + ".meta.json"
                    try:
                        with open(meta_path, "w", encoding="utf-8") as mf:
                            json.dump(meta, mf, ensure_ascii=False, indent=2)
                    except Exception:
                        # 元数据写入失败不影响主流程
                        pass
                    
                    # 将 CSV 数据写入 SQLite 数据库
                    if sqlite_conn is not None:
                        try:
                            table_name = safe_sheet
                            # 使用 pandas 将 DataFrame 写入 SQLite，列名对应 CSV 列名
                            df.to_sql(table_name, sqlite_conn, if_exists='replace', index=False)
                            
                            # 保存表格元数据到 _table_metadata
                            cursor = sqlite_conn.cursor()
                            cursor.execute("""
                                INSERT OR REPLACE INTO _table_metadata 
                                (table_name, sheet_name, description, row_count, column_count, csv_path, created_at)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (
                                table_name,
                                str(sheet_name),
                                meta['sheet_description'],
                                len(df),
                                len(df.columns),
                                os.path.abspath(csv_path),
                                ts
                            ))
                            
                            # 保存列元数据到 _column_metadata
                            for col_idx, col_meta in enumerate(meta["columns"]):
                                cursor.execute("""
                                    INSERT OR REPLACE INTO _column_metadata 
                                    (table_name, column_name, column_index, data_type, description, sample_values)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """, (
                                    table_name,
                                    col_meta['name'],
                                    col_idx,
                                    col_meta.get('dtype', 'TEXT'),
                                    col_meta['description'],
                                    json.dumps(col_meta.get('sample_values', []), ensure_ascii=False)
                                ))
                            sqlite_conn.commit()
                            logger.info(f"表 {table_name} 已写入 SQLite")
                        except Exception as e:
                            logger.warning(f"写入表 {safe_sheet} 到 SQLite 失败: {e}")
                    
                    output_files.append(os.path.abspath(csv_path))

                # 关闭 SQLite 连接
                if sqlite_conn is not None:
                    try:
                        sqlite_conn.close()
                        logger.info(f"SQLite 数据库已保存: {sqlite_path}")
                    except Exception:
                        pass
                
                # 写整本 Excel 的目录文件，便于大模型/检索使用
                try:
                    catalog_path = os.path.join(out_dir, f"{safe_base}_catalog.md")
                    with open(catalog_path, "w", encoding="utf-8") as cf:
                        cf.write("".join(catalog_lines))
                except Exception:
                    pass
                return output_files
            except Exception:
                # 没装 pandas / 读失败则回退
                if sqlite_conn is not None:
                    try:
                        sqlite_conn.close()
                    except Exception:
                        pass
                pass

        # 方案B：回退 openpyxl（仅支持 xlsx），同样生成基础元数据与整本目录
        if ext != ".xlsx" or load_workbook is None:
            if sqlite_conn is not None:
                try:
                    sqlite_conn.close()
                except Exception:
                    pass
            return output_files
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            catalog_lines.append(f"- Sheet 总数: {len(wb.worksheets)}\n\n")
            for ws in wb.worksheets:
                sheet_name = ws.title or "Sheet"
                safe_sheet = re.sub(r"[^\w\u4e00-\u9fa5]", "_", str(sheet_name))[:60] or "Sheet"
                csv_path = os.path.join(out_dir, f"{safe_base}__{safe_sheet}.csv")
                header = None
                rows_for_meta = []
                all_rows = []  # 用于写入 SQLite
                with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    for idx, row in enumerate(ws.iter_rows(values_only=True)):
                        row_values = ["" if v is None else v for v in row]
                        writer.writerow(row_values)
                        if idx == 0:
                            header = [str(v) if v is not None else "" for v in row]
                        else:
                            all_rows.append(row_values)
                        if len(rows_for_meta) < 10:
                            rows_for_meta.append(row_values)
                # 基于首行/前几行构造元数据
                columns_meta = []
                if header:
                    catalog_lines.append(f"## Sheet: {sheet_name}\n\n")
                    catalog_lines.append(f"- 描述: 工作表「{sheet_name}」，包含 {len(header)} 列。\n")
                    catalog_lines.append(f"- 对应 CSV: {os.path.abspath(csv_path)}\n\n")
                    catalog_lines.append("### 列信息\n\n")
                    for col_idx, col_name in enumerate(header):
                        # 收集该列的示例值
                        sample_values = []
                        for r in rows_for_meta[1:]:
                            if col_idx < len(r) and r[col_idx] not in ("", None):
                                sample_values.append(str(r[col_idx]))
                                if len(sample_values) >= 5:
                                    break
                        col_meta = {
                            "name": col_name or f"col_{col_idx+1}",
                            "description": f"列「{col_name or f'col_{col_idx+1}'}」，示例值: {', '.join(sample_values[:3])}" if sample_values else f"列「{col_name or f'col_{col_idx+1}'}」。",
                            "sample_values": sample_values,
                        }
                        columns_meta.append(col_meta)
                        samples_show = f" 示例值: {', '.join(sample_values[:3])}" if sample_values else ""
                        catalog_lines.append(f"- {col_meta['name']}: {col_meta['description']}{samples_show}\n")
                    catalog_lines.append("\n")
                meta = {
                    "sheet_name": str(sheet_name),
                    "sheet_description": f"工作表「{sheet_name}」，包含 {len(columns_meta)} 列。",
                    "columns": columns_meta,
                }
                meta_path = csv_path + ".meta.json"
                try:
                    with open(meta_path, "w", encoding="utf-8") as mf:
                        json.dump(meta, mf, ensure_ascii=False, indent=2)
                except Exception:
                    pass
                
                # 将数据写入 SQLite（openpyxl 方案）
                if sqlite_conn is not None and header:
                    try:
                        table_name = safe_sheet
                        cursor = sqlite_conn.cursor()
                        
                        # 清理列名，确保是有效的 SQL 列名
                        clean_columns = []
                        for col in header:
                            clean_col = re.sub(r"[^\w\u4e00-\u9fa5]", "_", str(col)) if col else f"col_{len(clean_columns)+1}"
                            clean_columns.append(clean_col)
                        
                        # 创建表，列名对应 CSV 列名
                        columns_def = ", ".join([f'"{col}" TEXT' for col in clean_columns])
                        cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                        cursor.execute(f'CREATE TABLE "{table_name}" ({columns_def})')
                        
                        # 插入数据
                        placeholders = ", ".join(["?" for _ in clean_columns])
                        for row in all_rows:
                            row_data = [str(v) if v not in ("", None) else "" for v in row[:len(clean_columns)]]
                            # 补齐列数
                            while len(row_data) < len(clean_columns):
                                row_data.append("")
                            cursor.execute(f'INSERT INTO "{table_name}" VALUES ({placeholders})', row_data)
                        
                        # 保存表格元数据
                        cursor.execute("""
                            INSERT OR REPLACE INTO _table_metadata 
                            (table_name, sheet_name, description, row_count, column_count, csv_path, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (
                            table_name,
                            str(sheet_name),
                            meta['sheet_description'],
                            len(all_rows),
                            len(columns_meta),
                            os.path.abspath(csv_path),
                            ts
                        ))
                        
                        # 保存列元数据
                        for col_idx, col_meta in enumerate(columns_meta):
                            cursor.execute("""
                                INSERT OR REPLACE INTO _column_metadata 
                                (table_name, column_name, column_index, data_type, description, sample_values)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (
                                table_name,
                                col_meta['name'],
                                col_idx,
                                'TEXT',
                                col_meta['description'],
                                json.dumps(col_meta.get('sample_values', []), ensure_ascii=False)
                            ))
                        sqlite_conn.commit()
                        logger.info(f"表 {table_name} 已写入 SQLite (openpyxl)")
                    except Exception as e:
                        logger.warning(f"写入表 {safe_sheet} 到 SQLite 失败 (openpyxl): {e}")
                
                output_files.append(os.path.abspath(csv_path))
            try:
                wb.close()
            except Exception:
                pass
            
            # 关闭 SQLite 连接
            if sqlite_conn is not None:
                try:
                    sqlite_conn.close()
                    logger.info(f"SQLite 数据库已保存: {sqlite_path}")
                except Exception:
                    pass
            
            # 写整本 Excel 的目录文件
            try:
                catalog_path = os.path.join(out_dir, f"{safe_base}_catalog.md")
                with open(catalog_path, "w", encoding="utf-8") as cf:
                    cf.write("".join(catalog_lines))
            except Exception:
                pass
            return output_files
        except Exception:
            if sqlite_conn is not None:
                try:
                    sqlite_conn.close()
                except Exception:
                    pass
            return output_files


    def read_pdf_basic(self, file_path):
        """
        初级PDF解析（使用本地库）
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            tuple: (query, content) 其中query通常为空字符串，content是提取的文本
        """
        try:
            # 尝试使用 PyMuPDF (fitz)
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(file_path)
                text_parts = []
                for page in doc:
                    text_parts.append(page.get_text())
                doc.close()
                content = '\n'.join(text_parts)
                return "", content
            except ImportError:
                pass
            
            # 尝试使用 PyPDF2
            try:
                import PyPDF2
                with open(file_path, 'rb') as f:
                    pdf_reader = PyPDF2.PdfReader(f)
                    text_parts = []
                    for page in pdf_reader.pages:
                        text_parts.append(page.extract_text())
                    content = '\n'.join(text_parts)
                    return "", content
            except ImportError:
                pass
            
            # 如果都没有安装，返回错误
            raise Exception("PDF解析库未安装。请安装 PyMuPDF 或 PyPDF2: pip install PyMuPDF 或 pip install PyPDF2")
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"初级PDF解析失败: {e}")
            return None, None
    
    def read_txt(self, file):
        with open(file, "r", encoding='utf-8') as f:
            content = f.read()
            return content
        return None
    
    def down_file(self, url, file_name):
        """
        下载文件功能
        :param url: 文件下载地址
        :param file_name: 包含路径的完整文件名
        :return: 下载结果信息
        """
        try:
            # 确保文件路径存在
            file_dir = os.path.dirname(file_name)
            if file_dir and not os.path.exists(file_dir):
                os.makedirs(file_dir)
            
            # 下载文件
            urllib.request.urlretrieve(url, file_name)
            
            # 检查文件是否成功下载
            if os.path.exists(file_name):
                file_size = os.path.getsize(file_name)
                return {
                    "success": True,
                    "message": "文件下载成功",
                    "file_path": file_name,
                    "file_size": file_size
                }
            else:
                return {
                    "success": False,
                    "message": "文件下载失败，文件未创建"
                }
        except urllib.error.URLError as e:
            return {
                "success": False,
                "message": f"URL错误: {str(e)}"
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"文件下载失败: {str(e)}"
            }

    
    